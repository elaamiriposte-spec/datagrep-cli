"""I/O utilities for datagrep - file loading and format detection."""

import csv
import io
import json
import logging
import os
import sys
from typing import Any, Dict, List, TextIO, Tuple, Union

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def load_excel_records(file: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load records from Excel XLSX file.
    
    Args:
        file: Path to Excel file.
        
    Returns:
        Tuple of (records list, field names list).
        
    Raises:
        ImportError: If openpyxl is not installed.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl is required for Excel support. Install with: pip install openpyxl')
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers: List[str] = [str(cell.value) for cell in ws[1]]
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                record[headers[i]] = cell.value
        records.append(record)
    return records, headers


def open_input_file(path: str, encoding: str) -> Union[TextIO, io.StringIO]:
    """Open input file with proper encoding handling.
    
    Args:
        path: File path or '-' for stdin.
        encoding: Text encoding to use.
        
    Returns:
        File object or stdin.
    """
    if path == '-':
        return sys.stdin
    return open(path, 'r', encoding=encoding, newline='')


def check_file_size(path: str) -> None:
    """Check file size and warn if it exceeds recommended limits.
    
    Args:
        path: File path to check.
    """
    if path == '-':
        return  # Can't check stdin
    
    try:
        file_size_mb: float = os.path.getsize(path) / (1024 * 1024)
        
        if file_size_mb > 1000:  # > 1GB
            logging.warning(
                "File size %.1fMB exceeds recommended limit (1GB). "
                "Performance may degrade or cause memory issues. "
                "See PERFORMANCE.md for optimization strategies.",
                file_size_mb
            )
        elif file_size_mb > 500:  # > 500MB
            logging.info(
                "Processing %.1fMB file. Consider using --limit or --where "
                "filters for better performance.",
                file_size_mb
            )
    except (OSError, IOError):
        pass  # File doesn't exist yet or can't access


def load_json_records(file: Union[TextIO, io.StringIO]) -> List[Dict[str, Any]]:
    """Load records from JSON file (array or newline-delimited).
    
    Args:
        file: File object containing JSON data.
        
    Returns:
        List of record dictionaries.
        
    Raises:
        ValueError: If JSON format is invalid.
    """
    try:
        data: Any = json.load(file)
    except json.JSONDecodeError:
        try:
            file.seek(0)
        except (OSError, AttributeError):
            text: str = file.read()
            file = io.StringIO(text)
            file.seek(0)
        records: List[Dict[str, Any]] = []
        for idx, line in enumerate(file, 1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f'Invalid JSON on line {idx}: {exc}')
        if not records:
            raise ValueError('No JSON objects found in input.')
        return records

    if isinstance(data, dict):
        raise ValueError('JSON input must be an array of objects or newline-delimited objects.')
    if not isinstance(data, list):
        raise ValueError('JSON input must be an array of objects.')
    if any(not isinstance(item, dict) for item in data):
        raise ValueError('JSON array must contain objects.')
    return data


def load_config(config_file: str) -> Dict[str, Any]:
    """Load configuration from JSON file.
    
    Args:
        config_file: Path to configuration JSON file.
        
    Returns:
        Dictionary of configuration options.
        
    Raises:
        FileNotFoundError: If config file doesn't exist.
        json.JSONDecodeError: If config file is invalid JSON.
    """
    with open(config_file, 'r', encoding='utf-8') as f:
        return json.load(f)
