"""datagrep - powerful command-line data search tool for CSV, JSON, and Excel files."""

import argparse
import csv
import io
import json
import logging
import operator
import os
import re
import sys
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, TextIO, Tuple, Union
from collections.abc import Iterator

__version__ = "1.0.0"

try:
    import colorama
    from colorama import Fore, Style
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False

try:
    import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataGrepError(Exception):
    """Custom exception for datagrep-cli errors."""
    pass


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.
    
    Supports both styles:
    - Legacy positional: datagrep file.csv columns value
    - Modern flags: datagrep --file file.csv --columns columns --search value
    
    Returns:
        Namespace with parsed arguments.
    """
    parser = argparse.ArgumentParser(
        prog='datagrep',
        description='Search CSV, JSON, or Excel records by field values with flexible matching modes.',
        epilog='Use - for input_file to read from stdin. Inspection modes do not require search parameters. Supports both positional args (legacy) and explicit flags (modern).',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        add_help=True
    )
    
    parser.add_argument(
        '--version', action='version', version=f'%(prog)s {__version__}',
        help='Show version and exit.'
    )
    
    # Legacy positional arguments (optional, for backward compatibility)
    parser.add_argument('input_file', nargs='?', default=None, help='(Legacy positional) Input file path or - for stdin.')
    parser.add_argument('columns', nargs='?', default=None, help='(Legacy positional) Comma-separated field names to search.')
    parser.add_argument('value', nargs='?', default=None, help='(Legacy positional) Search value or pattern.')
    
    # Modern explicit flags (optional, take precedence over positional)
    parser.add_argument(
        '--file', '-f', dest='file_flag', default=None,
        help='Input file path or - for stdin (modern flag style).'
    )
    parser.add_argument(
        '--columns', dest='columns_flag', default=None,
        help='Comma-separated field names to search (modern flag style).'
    )
    parser.add_argument(
        '--search', '-S', dest='search_flag', default=None,
        help='Search value or pattern (modern flag style).'
    )
    
    parser.add_argument(
        '--input-format', choices=['auto', 'csv', 'json', 'xlsx'], default='auto',
        help='Input format: auto-detect (default), csv, json, or xlsx.'
    )
    parser.add_argument(
        '--mode', choices=['contains', 'exact', 'startswith', 'endswith', 'regex'],
        default='contains',
        help='Search mode: contains (default), exact, startswith, endswith, or regex.'
    )
    parser.add_argument(
        '--ignore-case', '-i', action='store_true', help='Ignore case while searching.'
    )
    parser.add_argument(
        '--delimiter', '-d', default=',', help='CSV delimiter for input/output (default: ,).'
    )
    parser.add_argument(
        '--encoding', default='utf-8', help='Input/output file encoding (default: utf-8).'
    )
    parser.add_argument(
        '--output-format', choices=['csv', 'json', 'table', 'raw'], default='csv',
        help='Output format: csv (default), json, table, or raw.'
    )
    parser.add_argument(
        '--select', '-s', default='*',
        help='Comma-separated fields to include in output (default: all).'
    )
    parser.add_argument(
        '--output', '-o', help='Write results to file instead of stdout.'
    )
    parser.add_argument(
        '--limit', '-n', type=int, default=0,
        help='Maximum matching rows to print (0=unlimited, default).'
    )
    parser.add_argument(
        '--sort', help='Sort by column: "name:asc" or "name:desc" before searching.'
    )
    parser.add_argument(
        '--where', help='Filter records with condition: "column op value" (or/and supported).'
    )
    parser.add_argument(
        '--empty', action='store_true', help='Show only rows where specified column is empty.'
    )
    parser.add_argument(
        '--not-empty', action='store_true', help='Show only rows where specified column has a value.'
    )
    parser.add_argument(
        '--count', action='store_true', help='Only count matching records (mutually exclusive with other inspection modes).'
    )
    parser.add_argument(
        '--config', help='Load configuration from JSON file.'
    )
    parser.add_argument(
        '--color', action='store_true', help='Colorize output (requires colorama).'
    )
    parser.add_argument(
        '--progress', action='store_true', help='Show progress bar for large files (requires tqdm).'
    )
    parser.add_argument(
        '--preview', type=int, default=0, help='Show first N matching rows only (0=unlimited).'
    )
    parser.add_argument(
        '--sample', type=int, default=0, help='Show first N rows as sample (inspection mode).'
    )
    parser.add_argument(
        '--verbose', '-v', action='store_true', help='Enable verbose logging.'
    )
    parser.add_argument(
        '--debug', action='store_true', help='Enable debug logging.'
    )
    parser.add_argument(
        '--describe', action='store_true', help='Show file schema and sample data (inspection mode).'
    )
    
    args = parser.parse_args()
    return _reconcile_args(args)


def _reconcile_args(args: argparse.Namespace) -> argparse.Namespace:
    """Reconcile positional and flag-based arguments for backward compatibility.
    
    Priority: flags > positional arguments
    
    Args:
        args: Parsed arguments with both positional and flag versions.
        
    Returns:
        Reconciled arguments with consolidated values.
    """
    # Reconcile input_file / --file flag
    args.input_file = args.file_flag or args.input_file or '-'
    
    # Reconcile columns / --columns flag
    if args.columns_flag:
        args.columns = args.columns_flag
    
    # Reconcile value / --search flag
    if args.search_flag:
        args.value = args.search_flag
    
    # Remove flag attributes to avoid confusion
    delattr(args, 'file_flag')
    delattr(args, 'columns_flag')
    delattr(args, 'search_flag')
    
    return args


def validate_args(args: argparse.Namespace) -> argparse.Namespace:
    """Validate argument combinations and constraints.
    
    Args:
        args: Parsed arguments to validate.
        
    Returns:
        Validated arguments namespace.
        
    Raises:
        DataGrepError: When invalid argument combinations are detected.
    """
    inspection_modes: List[bool] = [args.count, args.describe, args.sample > 0, args.preview > 0]
    inspection_count: int = sum(inspection_modes)
    
    if inspection_count > 1:
        raise DataGrepError('Cannot combine --count, --describe, --sample, and --preview. Use only one.')
    
    in_inspection_mode: bool = inspection_count > 0
    
    if in_inspection_mode and args.value is not None:
        raise DataGrepError('Inspection modes (--count, --describe, --sample, --preview) cannot be used with search value.')
    
    # Validate --empty and --not-empty
    if args.empty and args.not_empty:
        raise DataGrepError(
            'Error: Cannot use --empty and --not-empty together.\n'
            '  Use only one filter at a time.\n'
            '  Examples:\n'
            '    datagrep file.csv phone --empty\n'
            '    datagrep file.csv email --not-empty'
        )
    
    if (args.empty or args.not_empty) and args.value is not None:
        raise DataGrepError(
            'Error: --empty and --not-empty filters do not take a search value.\n'
            '  These filters check if columns are empty/non-empty, not for specific values.\n'
            '  Incorrect: datagrep file.csv phone "123" --empty ❌\n'
            '  Correct:   datagrep file.csv phone --empty ✓'
        )
    
    if (args.empty or args.not_empty) and not args.columns:
        raise DataGrepError(
            'Error: --empty and --not-empty require a column name.\n'
            '  Specify which column to check.\n'
            '  Examples:\n'
            '    datagrep file.csv phone --empty\n'
            '    datagrep file.csv "email,phone" --empty'
        )
    
    if (args.empty or args.not_empty) and (args.where or args.sort):
        raise DataGrepError(
            'Error: Cannot combine --empty/--not-empty with --where or --sort.\n'
            '  These filters are mutually exclusive.\n'
            '  Choose one approach:\n'
            '    datagrep file.csv phone --empty\n'
            '    datagrep file.csv name john --where "age > 25"'
        )
    
    # Validate --where condition format if provided
    if args.where:
        try:
            # Try parsing the WHERE condition to catch format errors early
            parse_where_condition(args.where)
        except DataGrepError as e:
            raise DataGrepError(
                f'Error in --where condition: {str(e)}\n'
                '  Format: "column operator value"\n'
                '  Operators: ==, !=, >, <, >=, <=, contains, startswith, endswith\n'
                '  Logic: AND, OR\n'
                '  Examples:\n'
                '    --where "age > 25"\n'
                '    --where "status == active"\n'
                '    --where "status == active and age > 25"'
            )
    
    # Validate --sort format if provided
    if args.sort:
        if ':' not in args.sort:
            raise DataGrepError(
                'Error: Invalid --sort format.\n'
                '  Format: "column:asc" or "column:desc"\n'
                '  Examples:\n'
                '    --sort name:asc\n'
                '    --sort age:desc'
            )
        col, order = args.sort.split(':')
        if order.lower() not in ('asc', 'desc'):
            raise DataGrepError(
                f'Error: Invalid sort order "{order}".\n'
                '  Use "asc" (ascending) or "desc" (descending).\n'
                '  Examples:\n'
                '    --sort name:asc\n'
                '    --sort age:desc'
            )
    
    # Note: --where and --sort can be used without search value
    # They pre-filter/sort the records, then return all matches
    # Only non-inspection modes without filters require a search value
    
    if args.value and not args.columns:
        args.columns = '*'
    
    return args


def load_config(config_file: str) -> Dict[str, Any]:
    """Load configuration from JSON file.
    
    Args:
        config_file: Path to configuration JSON file.
        
    Returns:
        Dictionary of configuration options.
        
    Raises:
        FileNotFoundError: If config file doesn't exist.
        json.JSONDecodeError: If config file is invalid JSON.
    """
    with open(config_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_where_condition(condition: str) -> Callable[[Dict[str, Any]], bool]:
    """Parse WHERE condition with AND/OR logic.
    
    Args:
        condition: Filter condition string like "name == John" or "age > 18 and city contains NYC".
        
    Returns:
        Function that evaluates condition on a row dictionary.
        
    Raises:
        DataGrepError: If condition format is invalid.
    """
    def parse_single(cond: str) -> Callable[[Dict[str, Any]], bool]:
        """Parse single condition."""
        parts: List[str] = cond.strip().split()
        if len(parts) != 3:
            raise DataGrepError(
                f'Error in WHERE condition: "{cond}"\n'
                f'  Expected format: "column operator value"\n'
                f'  Make sure to use spaces around the operator.\n'
                f'  Examples:\n'
                f'    --where "name == john"\n'
                f'    --where "age > 25"\n'
                f'    --where "status == active"'
            )
        col, op, val = parts
        ops: Dict[str, Callable[[Any, Any], bool]] = {
            '==': operator.eq,
            '!=': operator.ne,
            '>': operator.gt,
            '<': operator.lt,
            '>=': operator.ge,
            '<=': operator.le,
            'contains': lambda a, b: b in a,
            'startswith': lambda a, b: a.startswith(b),
            'endswith': lambda a, b: a.endswith(b),
        }
        if op not in ops:
            raise DataGrepError(
                f'Error: Unknown operator "{op}" in condition "{cond}"\n'
                f'  Valid operators are: ==, !=, >, <, >=, <=, contains, startswith, endswith\n'
                f'  Examples:\n'
                f'    --where "status == active"\n'
                f'    --where "age > 25"\n'
                f'    --where "name contains john"'
            )
        return lambda row: ops[op](str(row.get(col, '')), val)

    if ' and ' in condition:
        conditions: List[str] = condition.split(' and ')
        funcs: List[Callable[[Dict[str, Any]], bool]] = [parse_single(c) for c in conditions]
        return lambda row: all(f(row) for f in funcs)
    elif ' or ' in condition:
        conditions = condition.split(' or ')
        funcs = [parse_single(c) for c in conditions]
        return lambda row: any(f(row) for f in funcs)
    else:
        return parse_single(condition)


def load_excel_records(file: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load records from Excel XLSX file.
    
    Args:
        file: Path to Excel file.
        
    Returns:
        Tuple of (records list, field names list).
        
    Raises:
        ImportError: If openpyxl is not installed.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl is required for Excel support. Install with: pip install openpyxl')
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers: List[str] = [str(cell.value) for cell in ws[1]]
    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        record: Dict[str, Any] = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                record[headers[i]] = cell.value
        records.append(record)
    return records, headers


def open_input_file(path: str, encoding: str) -> Union[TextIO, io.StringIO]:
    """Open input file with proper encoding handling.
    
    Args:
        path: File path or '-' for stdin.
        encoding: Text encoding to use.
        
    Returns:
        File object or stdin.
    """
    if path == '-':
        return sys.stdin
    return open(path, 'r', encoding=encoding, newline='')


def check_file_size(path: str) -> None:
    """Check file size and warn if it exceeds recommended limits.
    
    Args:
        path: File path to check.
    """
    if path == '-':
        return  # Can't check stdin
    
    try:
        file_size_mb: float = os.path.getsize(path) / (1024 * 1024)
        
        if file_size_mb > 1000:  # > 1GB
            logging.warning(
                "File size %.1fMB exceeds recommended limit (1GB). "
                "Performance may degrade or cause memory issues. "
                "See PERFORMANCE.md for optimization strategies.",
                file_size_mb
            )
        elif file_size_mb > 500:  # > 500MB
            logging.info(
                "Processing %.1fMB file. Consider using --limit or --where "
                "filters for better performance.",
                file_size_mb
            )
    except (OSError, IOError):
        pass  # File doesn't exist yet or can't access


def build_matcher(value: str, mode: str, ignore_case: bool) -> Callable[[str], bool]:
    """Build a matcher function based on search mode.
    
    Args:
        value: Search string or pattern.
        mode: Search mode (contains, exact, startswith, endswith, regex).
        ignore_case: Whether to ignore case.
        
    Returns:
        Function that tests if text matches criteria.
        
    Raises:
        ValueError: If mode is unknown.
    """
    if ignore_case:
        value = value.lower()

    if mode == 'contains':
        return lambda text: value in text.lower() if ignore_case else value in text
    if mode == 'exact':
        return lambda text: text.lower() == value if ignore_case else text == value
    if mode == 'startswith':
        return lambda text: text.lower().startswith(value) if ignore_case else text.startswith(value)
    if mode == 'endswith':
        return lambda text: text.lower().endswith(value) if ignore_case else text.endswith(value)
    if mode == 'regex':
        flags: int = re.IGNORECASE if ignore_case else 0
        pattern: re.Pattern[str] = re.compile(value, flags)
        return lambda text: bool(pattern.search(text))
    raise ValueError(f'Unknown mode: {mode}')


def load_json_records(file: Union[TextIO, io.StringIO]) -> List[Dict[str, Any]]:
    """Load records from JSON file (array or newline-delimited).
    
    Args:
        file: File object containing JSON data.
        
    Returns:
        List of record dictionaries.
        
    Raises:
        ValueError: If JSON format is invalid.
    """
    try:
        data: Any = json.load(file)
    except json.JSONDecodeError:
        try:
            file.seek(0)
        except (OSError, AttributeError):
            text: str = file.read()
            file = io.StringIO(text)
            file.seek(0)
        records: List[Dict[str, Any]] = []
        for idx, line in enumerate(file, 1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f'Invalid JSON on line {idx}: {exc}')
        if not records:
            raise ValueError('No JSON objects found in input.')
        return records

    if isinstance(data, dict):
        raise ValueError('JSON input must be an array of objects or newline-delimited objects.')
    if not isinstance(data, list):
        raise ValueError('JSON input must be an array of objects.')
    if any(not isinstance(item, dict) for item in data):
        raise ValueError('JSON array must contain objects.')
    return data


def format_table(rows: List[Dict[str, Any]], fieldnames: List[str], color: bool = False) -> str:
    """Format records as aligned ASCII table.
    
    Args:
        rows: List of record dictionaries.
        fieldnames: Column names to include.
        color: Whether to use ANSI color codes.
        
    Returns:
        Formatted table string.
    """
    if not rows:
        return ''

    widths: Dict[str, int] = {field: len(field) for field in fieldnames}
    for row in rows:
        for field in fieldnames:
            widths[field] = max(widths[field], len(str(row.get(field, ''))))

    header_parts: List[str] = [field.ljust(widths[field]) for field in fieldnames]
    if color and COLORAMA_AVAILABLE:
        header: str = Fore.CYAN + ' | '.join(header_parts) + Style.RESET_ALL
        separator: str = Fore.YELLOW + '-+-'.join('-' * widths[field] for field in fieldnames) + Style.RESET_ALL
    else:
        header = ' | '.join(header_parts)
        separator = '-+-'.join('-' * widths[field] for field in fieldnames)
    lines: List[str] = [header, separator]
    for row in rows:
        line_parts = [str(row.get(field, '')).ljust(widths[field]) for field in fieldnames]
        lines.append(' | '.join(line_parts))
    return '\n'.join(lines)


def _should_load_eagerly(args: argparse.Namespace) -> bool:
    """Determine if records should be loaded eagerly (all at once) vs lazily.
    
    Eager loading is required for:
    - Inspection modes (--describe, --count without value, --sample, --preview without value)
    - Filtering (--where, --sort, --empty, --not-empty)
    - When no limit is specified
    
    Args:
        args: Parsed command-line arguments.
        
    Returns:
        True if should load all records eagerly, False for lazy loading.
    """
    inspection_modes: List[bool] = [args.describe, args.sample > 0, args.preview > 0]
    has_inspection = any(inspection_modes)
    has_filters = bool(args.where or args.sort or args.empty or args.not_empty)
    has_no_value = args.value is None
    
    # Need eager loading if:
    if has_no_value:  # Inspection mode without value
        return True
    if has_inspection:  # Any inspection mode
        return True
    if has_filters:  # Any filtering
        return True
    if not args.limit:  # No limit specified, need all for full output
        return True
    
    # Can use lazy loading: search with --limit, no filters, no inspection
    return False


def main() -> None:
    """Main entry point for datagrep CLI."""
    args: argparse.Namespace = parse_args()
    args = validate_args(args)

    if args.debug:
        logging.basicConfig(level=logging.DEBUG, format='%(levelname)s: %(message)s')
    elif args.verbose:
        logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    else:
        logging.basicConfig(level=logging.WARNING)

    # Load config if provided
    if args.config:
        config = load_config(args.config)
        for key, value in config.items():
            if hasattr(args, key):
                setattr(args, key, value)
    # Check file size before loading
    check_file_size(args.input_file)
    try:
        with open_input_file(args.input_file, args.encoding) as csvfile:
            input_format: str = args.input_format
            if input_format == 'auto':
                if args.input_file != '-' and args.input_file.lower().endswith('.json'):
                    input_format = 'json'
                elif args.input_file != '-' and args.input_file.lower().endswith('.xlsx'):
                    input_format = 'xlsx'
                elif args.input_file != '-' and args.input_file.lower().endswith('.csv'):
                    input_format = 'csv'
                elif args.input_file == '-':
                    text: str = csvfile.read()
                    csvfile = io.StringIO(text)
                    input_format = 'json' if text.lstrip().startswith(('{', '[')) else 'csv'
                else:
                    input_format = 'csv'

            if input_format == 'csv':
                reader = csv.DictReader(csvfile, delimiter=args.delimiter)
                if not reader.fieldnames:
                    raise ValueError('CSV input has no headers.')
                available_columns: List[str] = reader.fieldnames
                
                # Decide whether to load eagerly or lazily
                use_lazy: bool = not _should_load_eagerly(args)
                if use_lazy:
                    # Lazy loading: keep as iterator, don't load all records
                    records: Union[List[Dict[str, Any]], Iterator[Dict[str, Any]]] = reader
                    logging.debug("Using lazy loading mode for CSV search")
                else:
                    # Eager loading: load all records into memory
                    records = list(reader)
                    logging.debug("Using eager loading mode for CSV (loaded %d records)", len(records))
            elif input_format == 'xlsx':
                # Excel always loads eagerly (openpyxl limitation)
                records, available_columns = load_excel_records(args.input_file)
                logging.debug("Loaded %d records from Excel file", len(records))
            else:
                # JSON always loads eagerly
                records = load_json_records(csvfile)
                logging.debug("Loaded %d records from JSON file", len(records))
                seen = set()
                available_columns = []
                for row in records:
                    for key in row.keys():
                        if key not in seen:
                            seen.add(key)
                            available_columns.append(key)

            columns: List[str] = [col.strip() for col in args.columns.split(',') if col.strip()] if args.columns else list(available_columns)
            selected_columns: List[str] = [col.strip() for col in args.select.split(',') if col.strip()]

            # For lazy loading, we can't get len() until we consume the iterator
            # This is fine because lazy loading only applies to search with --limit
            if not isinstance(records, list):
                # Still an iterator (lazy loading mode)
                records_count: Optional[int] = None
                logging.debug("Records in lazy loading mode - count unavailable until search")
            else:
                records_count = len(records)
                logging.info("Loaded %d records with fields: %s", records_count, ', '.join(available_columns))

            if args.describe:
                print("Schema:")
                for col in available_columns:
                    print(f"  - {col}")
                return
            elif args.count and not args.value:
                if isinstance(records, list):
                    print(len(records))
                else:
                    # This shouldn't happen with eager loading check, but fallback to counting
                    records = list(records)
                    print(len(records))
                return
            elif args.sample:
                if isinstance(records, list):
                    sample_rows: List[Dict[str, Any]] = records[:args.sample]
                else:
                    # Convert to list and take sample
                    records = list(records)
                    sample_rows = records[:args.sample]
                print(format_table(sample_rows, available_columns, args.color))
                return
            elif args.preview and not args.value:
                sample_rows = records[:args.preview]
                print(format_table(sample_rows, available_columns, args.color))
                return

            if not args.value:
                # Handle --empty and --not-empty filters (no search value needed)
                if args.empty or args.not_empty:
                    filter_columns: List[str] = [col.strip() for col in args.columns.split(',') if col.strip()]
                    
                    if args.empty:
                        logging.debug("Filtering for empty values in columns: %s", ', '.join(filter_columns))
                        records = [r for r in records if any(str(r.get(col, '')).strip() == '' for col in filter_columns)]
                        logging.info("After empty filter: %d records", len(records))
                    else:  # --not-empty
                        logging.debug("Filtering for non-empty values in columns: %s", ', '.join(filter_columns))
                        records = [r for r in records if any(str(r.get(col, '')).strip() != '' for col in filter_columns)]
                        logging.info("After not-empty filter: %d records", len(records))
                    
                    if not records:
                        print("No records match the filter.")
                        return
                    
                    # Show filtered results
                    if selected_columns == ['*']:
                        selected_columns = available_columns
                    print(format_table(records, selected_columns, args.color))
                    return
                
                # Handle --where and --sort without search value
                if args.where or args.sort:
                    # Apply where filter
                    if args.where:
                        logging.debug("Applying where filter: %s", args.where)
                        where_func: Callable[[Dict[str, Any]], bool] = parse_where_condition(args.where)
                        records = [r for r in records if where_func(r)]
                        logging.info("After where filter: %d records", len(records))

                    # Apply sorting
                    if args.sort:
                        logging.debug("Applying sort: %s", args.sort)
                        sort_col, sort_order = args.sort.split(':')
                        reverse: bool = sort_order.lower() == 'desc'
                        records.sort(key=lambda r: str(r.get(sort_col, '')), reverse=reverse)
                        logging.info("Records sorted by %s %s", sort_col, sort_order)

                    if not records:
                        print("No records match the filter.")
                        return
                    
                    # Show filtered results
                    if selected_columns == ['*']:
                        selected_columns = available_columns
                    print(format_table(records, selected_columns, args.color))
                    return
                
                # Show schema and sample when no search value provided and no filters
                print("Schema:")
                for col in available_columns:
                    print(f"  - {col}")
                print("\nSample rows (first 10):")
                sample_rows = records[:10]
                if sample_rows:
                    print(format_table(sample_rows, available_columns, args.color))
                return

            # Search path: args.value is provided
            
            if columns == ['*']:
                columns = available_columns
            if selected_columns == ['*']:
                selected_columns = available_columns

            missing_search = [col for col in columns if col not in available_columns]
            if missing_search:
                raise ValueError(
                    f"Error: Search column(s) not found: {', '.join(missing_search)}\n"
                    f"  Available columns in this file: {', '.join(available_columns)}\n"
                    f"  Check the column names are spelled correctly (case-sensitive).\n"
                    f"  Use --inspect to see all available columns."
                )

            missing_select = [col for col in selected_columns if col not in available_columns]
            if missing_select:
                raise ValueError(
                    f"Error: --select column(s) not found: {', '.join(missing_select)}\n"
                    f"  Available columns in this file: {', '.join(available_columns)}\n"
                    f"  Check the column names are spelled correctly (case-sensitive).\n"
                    f"  Use --inspect to see all available columns."
                )

            # Apply where filter (when search value is provided)
            if args.where:
                logging.debug("Applying where filter: %s", args.where)
                where_func: Callable[[Dict[str, Any]], bool] = parse_where_condition(args.where)
                records = [r for r in records if where_func(r)]
                logging.info("After where filter: %d records", len(records))

            # Apply sorting (when search value is provided)
            if args.sort:
                logging.debug("Applying sort: %s", args.sort)
                sort_col, sort_order = args.sort.split(':')
                reverse: bool = sort_order.lower() == 'desc'
                records.sort(key=lambda r: str(r.get(sort_col, '')), reverse=reverse)
                logging.info("Records sorted by %s %s", sort_col, sort_order)

            matcher: Callable[[str], bool] = build_matcher(args.value, args.mode, args.ignore_case)
            logging.debug("Searching for '%s' in columns: %s", args.value, ', '.join(columns))

            if args.count:
                # Count only
                count: int = 0
                iterator: Union[Iterator[Dict[str, Any]], List[Dict[str, Any]]] = (
                    tqdm.tqdm(records) if args.progress and TQDM_AVAILABLE and len(records) > 1000 else records
                )
                for row in iterator:
                    if any(matcher(str(row.get(col, ''))) for col in columns):
                        count += 1
                        if args.limit and count >= args.limit:
                            break
                print(count)
                logging.info("Found %d matching records", count)
                return

            matches: List[Dict[str, Any]] = []
            count = 0
            iterator = tqdm.tqdm(records) if args.progress and TQDM_AVAILABLE and len(records) > 1000 else records
            for row in iterator:
                if any(matcher(str(row.get(col, ''))) for col in columns):
                    matches.append(row)
                    count += 1
                    if args.limit and count >= args.limit:
                        break

            if args.preview and matches:
                matches = matches[:args.preview]
                logging.info("Preview limited to first %d matches", args.preview)

            logging.info("Found %d matching records", len(matches))

            output_file: Union[TextIO, io.TextIOBase] = open(args.output, 'w', encoding=args.encoding, newline='') if args.output else sys.stdout
            try:
                if args.output_format == 'csv':
                    writer = csv.DictWriter(output_file, fieldnames=selected_columns, delimiter=args.delimiter)
                    writer.writeheader()
                    for row in matches:
                        writer.writerow({col: row.get(col, '') for col in selected_columns})
                elif args.output_format == 'json':
                    json.dump(
                        [{col: row.get(col, '') for col in selected_columns} for row in matches],
                        output_file,
                        ensure_ascii=False,
                        indent=2
                    )
                    output_file.write('\n')
                elif args.output_format == 'table':
                    output_file.write(format_table(matches, selected_columns, args.color))
                    output_file.write('\n')
                else:
                    for row in matches:
                        subset = {col: row.get(col, '') for col in selected_columns}
                        output_file.write(f"{subset}\n")
            finally:
                if args.output:
                    output_file.close()

            if count == 0:
                print(
                    f"No records found where any of {', '.join(columns)} {args.mode} '{args.value}'",
                    file=sys.stderr
                )
    except FileNotFoundError:
        print(
            f"Error: File not found: '{args.input_file}'\n"
            f"  The file does not exist at that path.\n"
            f"  Check the path and file name (case-sensitive on Linux/Mac).\n"
            f"  Try using: pwd (to see current directory) and ls/dir (to list files)",
            file=sys.stderr
        )
        sys.exit(1)
    except UnicodeDecodeError:
        print(
            f"Error: Failed to decode '{args.input_file}' with encoding {args.encoding}.\n"
            f"  This file may be encoded differently (e.g., UTF-8 vs Latin-1).\n"
            f"  Try using a different encoding with: --encoding utf-8 or --encoding latin-1\n"
            f"  To detect encoding: file {args.input_file}",
            file=sys.stderr
        )
        sys.exit(1)
    except (ValueError, DataGrepError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)
    except re.error as exc:
        print(
            f"Regex error: Invalid regular expression pattern\n"
            f"  Error details: {exc}\n"
            f"  Check your --search value for special regex characters or syntax errors.\n"
            f"  If using --mode regex, ensure the pattern is valid regex (not a simple string).\n"
            f"  Examples of valid regex patterns:\n"
            f"    --mode regex --search '^[A-Z]'          (starts with capital letter)\n"
            f"    --mode regex --search '[0-9]{{3}}-[0-9]{{4}}' (phone number pattern)",
            file=sys.stderr
        )
        sys.exit(1)
    except Exception as exc:
        print(f"An unexpected error occurred: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
