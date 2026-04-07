import argparse
import csv
import io
import json
import operator
import os
import re
import sys
from pathlib import Path

try:
    import colorama
    from colorama import Fore, Style
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False

try:
    import tqdm
    TQDM_AVAILABLE = True
except ImportError:
    TQDM_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def parse_args():
    parser = argparse.ArgumentParser(
        description='Search CSV, JSON, or Excel records by field values with flexible matching modes.',
        epilog='Use - for csv_file to read from stdin.'
    )
    parser.add_argument('csv_file', nargs='?', default='-', help='Input file path or - for stdin.')
    parser.add_argument('columns', help='Comma-separated field names to search.')
    parser.add_argument('value', help='Search value or pattern.')
    parser.add_argument(
        '--input-format', choices=['auto', 'csv', 'json', 'xlsx'], default='auto',
        help='Input format to read: csv, json, xlsx, or auto-detect from file extension/content.'
    )
    parser.add_argument(
        '--mode', choices=['contains', 'exact', 'startswith', 'endswith', 'regex'],
        default='contains',
        help='Search mode to use (default: contains).'
    )
    parser.add_argument(
        '--ignore-case', '-i', action='store_true', help='Ignore case while searching.'
    )
    parser.add_argument(
        '--delimiter', '-d', default=',', help='CSV delimiter for input/output (default: ,).'
    )
    parser.add_argument(
        '--encoding', default='utf-8', help='Input/output file encoding (default: utf-8).'
    )
    parser.add_argument(
        '--output-format', choices=['csv', 'json', 'table', 'raw'], default='csv',
        help='Output format for matching rows (default: csv).'
    )
    parser.add_argument(
        '--select', '-s', default='*',
        help='Comma-separated fields to include in output (default: all fields).'
    )
    parser.add_argument(
        '--output', '-o', help='Write results to a file instead of stdout.'
    )
    parser.add_argument(
        '--limit', '-n', type=int, default=0,
        help='Maximum number of matching rows to print (0 = no limit).'
    )
    parser.add_argument(
        '--sort', help='Sort records by column:asc or column:desc before searching.'
    )
    parser.add_argument(
        '--where', help='Filter records with condition like "column > value" before searching.'
    )
    parser.add_argument(
        '--count', action='store_true', help='Only count matching records, do not print them.'
    )
    parser.add_argument(
        '--config', help='Load configuration from JSON file.'
    )
    parser.add_argument(
        '--color', action='store_true', help='Colorize output (requires colorama).'
    )
    parser.add_argument(
        '--progress', action='store_true', help='Show progress bar for large files (requires tqdm).'
    )
    return parser.parse_args()


def load_config(config_file):
    with open(config_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_where_condition(condition):
    parts = condition.split()
    if len(parts) != 3:
        raise ValueError('Where condition must be "column op value"')
    col, op, val = parts
    ops = {
        '==': operator.eq,
        '!=': operator.ne,
        '>': operator.gt,
        '<': operator.lt,
        '>=': operator.ge,
        '<=': operator.le,
        'contains': lambda a, b: b in a,
        'startswith': lambda a, b: a.startswith(b),
        'endswith': lambda a, b: a.endswith(b),
    }
    if op not in ops:
        raise ValueError(f'Unknown operator {op}')
    return lambda row: ops[op](str(row.get(col, '')), val)


def load_excel_records(file):
    if not OPENPYXL_AVAILABLE:
        raise ImportError('openpyxl is required for Excel support. Install with: pip install openpyxl')
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2):
        record = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                record[headers[i]] = cell.value
        records.append(record)
    return records, headers


def open_input_file(path, encoding):
    if path == '-':
        return sys.stdin
    return open(path, 'r', encoding=encoding, newline='')


def build_matcher(value, mode, ignore_case):
    if ignore_case:
        value = value.lower()

    if mode == 'contains':
        return lambda text: value in text.lower() if ignore_case else value in text
    if mode == 'exact':
        return lambda text: text.lower() == value if ignore_case else text == value
    if mode == 'startswith':
        return lambda text: text.lower().startswith(value) if ignore_case else text.startswith(value)
    if mode == 'endswith':
        return lambda text: text.lower().endswith(value) if ignore_case else text.endswith(value)
    if mode == 'regex':
        flags = re.IGNORECASE if ignore_case else 0
        pattern = re.compile(value, flags)
        return lambda text: bool(pattern.search(text))
    raise ValueError(f'Unknown mode: {mode}')


def load_json_records(file):
    try:
        data = json.load(file)
    except json.JSONDecodeError:
        try:
            file.seek(0)
        except (OSError, AttributeError):
            text = file.read()
            file = io.StringIO(text)
            file.seek(0)
        records = []
        for idx, line in enumerate(file, 1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f'Invalid JSON on line {idx}: {exc}')
        if not records:
            raise ValueError('No JSON objects found in input.')
        return records

    if isinstance(data, dict):
        raise ValueError('JSON input must be an array of objects or newline-delimited objects.')
    if not isinstance(data, list):
        raise ValueError('JSON input must be an array of objects.')
    if any(not isinstance(item, dict) for item in data):
        raise ValueError('JSON array must contain objects.')
    return data


def format_table(rows, fieldnames, color=False):
    if not rows:
        return ''

    widths = {field: len(field) for field in fieldnames}
    for row in rows:
        for field in fieldnames:
            widths[field] = max(widths[field], len(str(row.get(field, ''))))

    header_parts = [field.ljust(widths[field]) for field in fieldnames]
    if color and COLORAMA_AVAILABLE:
        header = Fore.CYAN + ' | '.join(header_parts) + Style.RESET_ALL
        separator = Fore.YELLOW + '-+-'.join('-' * widths[field] for field in fieldnames) + Style.RESET_ALL
    else:
        header = ' | '.join(header_parts)
        separator = '-+-'.join('-' * widths[field] for field in fieldnames)
    lines = [header, separator]
    for row in rows:
        line_parts = [str(row.get(field, '')).ljust(widths[field]) for field in fieldnames]
        lines.append(' | '.join(line_parts))
    return '\n'.join(lines)


def main():
    args = parse_args()

    # Load config if provided
    if args.config:
        config = load_config(args.config)
        for key, value in config.items():
            if hasattr(args, key):
                setattr(args, key, value)

    columns = [col.strip() for col in args.columns.split(',') if col.strip()]
    selected_columns = [col.strip() for col in args.select.split(',') if col.strip()]

    try:
        with open_input_file(args.csv_file, args.encoding) as csvfile:
            input_format = args.input_format
            if input_format == 'auto':
                if args.csv_file != '-' and args.csv_file.lower().endswith('.json'):
                    input_format = 'json'
                elif args.csv_file != '-' and args.csv_file.lower().endswith('.xlsx'):
                    input_format = 'xlsx'
                elif args.csv_file != '-' and args.csv_file.lower().endswith('.csv'):
                    input_format = 'csv'
                elif args.csv_file == '-':
                    text = csvfile.read()
                    csvfile = io.StringIO(text)
                    input_format = 'json' if text.lstrip().startswith(('{', '[')) else 'csv'
                else:
                    input_format = 'csv'

            if input_format == 'csv':
                reader = csv.DictReader(csvfile, delimiter=args.delimiter)
                if not reader.fieldnames:
                    raise ValueError('CSV input has no headers.')
                available_columns = reader.fieldnames
                records = list(reader)
            elif input_format == 'xlsx':
                records, available_columns = load_excel_records(args.csv_file)
            else:
                records = load_json_records(csvfile)
                seen = set()
                available_columns = []
                for row in records:
                    for key in row.keys():
                        if key not in seen:
                            seen.add(key)
                            available_columns.append(key)

            # Apply where filter
            if args.where:
                where_func = parse_where_condition(args.where)
                records = [r for r in records if where_func(r)]

            # Apply sorting
            if args.sort:
                sort_col, sort_order = args.sort.split(':')
                reverse = sort_order.lower() == 'desc'
                records.sort(key=lambda r: str(r.get(sort_col, '')), reverse=reverse)

            if columns == ['*']:
                columns = available_columns
            if selected_columns == ['*']:
                selected_columns = available_columns

            missing_search = [col for col in columns if col not in available_columns]
            if missing_search:
                raise ValueError(
                    f"Search field(s) not found: {', '.join(missing_search)}. "
                    f"Available fields: {', '.join(available_columns)}"
                )

            missing_select = [col for col in selected_columns if col not in available_columns]
            if missing_select:
                raise ValueError(
                    f"Selected field(s) not found: {', '.join(missing_select)}. "
                    f"Available fields: {', '.join(available_columns)}"
                )

            matcher = build_matcher(args.value, args.mode, args.ignore_case)

            if args.count:
                # Count only
                count = 0
                iterator = tqdm.tqdm(records) if args.progress and TQDM_AVAILABLE and len(records) > 1000 else records
                for row in iterator:
                    if any(matcher(str(row.get(col, ''))) for col in columns):
                        count += 1
                        if args.limit and count >= args.limit:
                            break
                print(count)
                return

            matches = []
            count = 0
            iterator = tqdm.tqdm(records) if args.progress and TQDM_AVAILABLE and len(records) > 1000 else records
            for row in iterator:
                if any(matcher(str(row.get(col, ''))) for col in columns):
                    matches.append(row)
                    count += 1
                    if args.limit and count >= args.limit:
                        break

            output_file = open(args.output, 'w', encoding=args.encoding, newline='') if args.output else sys.stdout
            try:
                if args.output_format == 'csv':
                    writer = csv.DictWriter(output_file, fieldnames=selected_columns, delimiter=args.delimiter)
                    writer.writeheader()
                    for row in matches:
                        writer.writerow({col: row.get(col, '') for col in selected_columns})
                elif args.output_format == 'json':
                    json.dump(
                        [{col: row.get(col, '') for col in selected_columns} for row in matches],
                        output_file,
                        ensure_ascii=False,
                        indent=2
                    )
                    output_file.write('\n')
                elif args.output_format == 'table':
                    output_file.write(format_table(matches, selected_columns, args.color))
                    output_file.write('\n')
                else:
                    for row in matches:
                        subset = {col: row.get(col, '') for col in selected_columns}
                        output_file.write(f"{subset}\n")
            finally:
                if args.output:
                    output_file.close()

            if count == 0:
                print(
                    f"No records found where any of {', '.join(columns)} {args.mode} '{args.value}'",
                    file=sys.stderr
                )
    except FileNotFoundError:
        print(f"Error: File '{args.csv_file}' not found.", file=sys.stderr)
        sys.exit(1)
    except UnicodeDecodeError:
        print(f"Error: Failed to decode '{args.csv_file}' with encoding {args.encoding}.", file=sys.stderr)
        sys.exit(1)
    except ValueError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)
    except re.error as exc:
        print(f"Regex error: {exc}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"An unexpected error occurred: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
